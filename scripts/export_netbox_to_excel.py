#!/usr/bin/env python3
"""Export NetBox source-of-truth data to a multi-sheet Excel workbook."""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INVENTORY = REPO_ROOT / "inventories" / "lab" / "netbox_inventory.yml"
DEFAULT_OUTPUT = REPO_ROOT / "generated_reports" / "NetBox_Master_Export.xlsx"


def load_project_netbox_defaults(path: Path = DEFAULT_INVENTORY) -> dict[str, str]:
    """Read simple api_endpoint/token defaults from the NetBox inventory file."""
    defaults: dict[str, str] = {}
    if not path.exists():
        return defaults

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or ":" not in line:
            continue

        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip().strip("'\"")
        if key == "api_endpoint":
            defaults["url"] = value
        elif key == "token":
            defaults["token"] = value

    return defaults


def import_dependencies() -> tuple[Any, Any, Any]:
    try:
        import pandas as pd
        import requests
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        missing = exc.name or "required package"
        print(
            f"Missing Python package: {missing}\n"
            "Install dependencies with: python3 -m pip install -r requirements.txt",
            file=sys.stderr,
        )
        raise SystemExit(2) from exc

    openpyxl_helpers = {
        "Font": Font,
        "PatternFill": PatternFill,
        "get_column_letter": get_column_letter,
    }
    return pd, requests, openpyxl_helpers


def build_headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Token {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def fetch_netbox_data(
    session: Any,
    base_url: str,
    endpoint: str,
    headers: dict[str, str],
    timeout: int,
    verify_tls: bool,
) -> list[dict[str, Any]]:
    """Fetch all objects from a NetBox API endpoint, following pagination."""
    api_root = base_url.rstrip("/") + "/api/"
    url = api_root + endpoint.strip("/")
    results: list[dict[str, Any]] = []

    while url:
        response = session.get(url, headers=headers, timeout=timeout, verify=verify_tls)
        response.raise_for_status()
        payload = response.json()

        if isinstance(payload, dict) and "results" in payload:
            results.extend(payload.get("results") or [])
            url = payload.get("next")
        elif isinstance(payload, list):
            results.extend(payload)
            url = None
        else:
            raise ValueError(f"Unexpected response format from {endpoint}")

    return results


def label(value: Any, *keys: str) -> str:
    """Return a human-friendly field from a NetBox nested object."""
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in keys or ("label", "name", "display", "value"):
            if value.get(key) not in (None, ""):
                return str(value[key])
        return ""
    return str(value)


def nested(data: dict[str, Any], *keys: str, default: str = "") -> str:
    value: Any = data
    for key in keys:
        if not isinstance(value, dict):
            return default
        value = value.get(key)
        if value is None:
            return default
    return label(value)


def vlan_ref(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, dict):
        vid = value.get("vid") or value.get("vlan_id")
        name = value.get("name") or value.get("display") or ""
        if vid and name:
            return f"{vid} - {name}"
        return str(vid or name or value.get("id") or "")
    return str(value)


def vlan_refs(values: Any) -> str:
    if not values:
        return ""
    if isinstance(values, list):
        return ", ".join(filter(None, (vlan_ref(item) for item in values)))
    return vlan_ref(values)


def endpoint_ref(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return ", ".join(filter(None, (endpoint_ref(item) for item in value)))
    if isinstance(value, dict):
        device = nested(value, "device")
        name = value.get("name") or value.get("display") or ""
        return f"{device} {name}".strip()
    return str(value)


def cable_peer(interface: dict[str, Any]) -> str:
    for key in ("link_peers", "connected_endpoints"):
        if interface.get(key):
            return endpoint_ref(interface[key])
    if interface.get("connected_endpoint"):
        return endpoint_ref(interface["connected_endpoint"])
    cable = interface.get("cable")
    if isinstance(cable, dict):
        return cable.get("label") or cable.get("display") or str(cable.get("id") or "")
    return ""


def interface_role(interface: dict[str, Any]) -> str:
    custom_fields = interface.get("custom_fields") or {}
    return str(
        interface.get("cf_port_role")
        or custom_fields.get("port_role")
        or custom_fields.get("cf_port_role")
        or ""
    )


def build_inventory(devices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for dev in devices:
        device_type = dev.get("device_type") or {}
        rows.append(
            {
                "Site": nested(dev, "site"),
                "Location": nested(dev, "location"),
                "Rack": nested(dev, "rack"),
                "DEVICE ROLE": nested(dev, "device_role"),
                "Manufacturer": nested(device_type, "manufacturer"),
                "MODEL": device_type.get("model", ""),
                "Platform": nested(dev, "platform"),
                "HOST NAME": dev.get("name", ""),
                "MGT IP": nested(dev, "primary_ip4", "address"),
                "STATUS": nested(dev, "status"),
                "Serial": dev.get("serial", ""),
            }
        )
    return rows


def build_vlans(vlans: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for vlan in vlans:
        rows.append(
            {
                "VLAN Number": vlan.get("vid", ""),
                "VLAN Name": vlan.get("name", ""),
                "Site": nested(vlan, "site"),
                "Group": nested(vlan, "group"),
                "Tenant": nested(vlan, "tenant"),
                "Trang thai": nested(vlan, "status"),
                "Mo ta": vlan.get("description", ""),
            }
        )
    return rows


def build_prefixes(prefixes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for prefix in prefixes:
        rows.append(
            {
                "Prefix": prefix.get("prefix", ""),
                "VLAN": vlan_ref(prefix.get("vlan")),
                "Site": nested(prefix, "site"),
                "VRF": nested(prefix, "vrf"),
                "Tenant": nested(prefix, "tenant"),
                "Status": nested(prefix, "status"),
                "Description": prefix.get("description", ""),
            }
        )
    return rows


def build_ports(interfaces: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for interface in interfaces:
        rows.append(
            {
                "Thiet bi": nested(interface, "device"),
                "Cong (Interface)": interface.get("name", ""),
                "Loai cong": nested(interface, "type"),
                "Mo ta": interface.get("description", ""),
                "Trang thai": "Enabled" if interface.get("enabled") else "Disabled",
                "Mode": nested(interface, "mode"),
                "Untagged VLAN": vlan_ref(interface.get("untagged_vlan")),
                "Tagged VLANs": vlan_refs(interface.get("tagged_vlans")),
                "Port Role": interface_role(interface),
                "Cable/Peer": cable_peer(interface),
            }
        )
    return rows


def build_ip_addresses(addresses: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for address in addresses:
        assigned = address.get("assigned_object") or {}
        rows.append(
            {
                "IP Address": address.get("address", ""),
                "DNS Name": address.get("dns_name", ""),
                "Device": nested(assigned, "device"),
                "Interface": assigned.get("name", "") if isinstance(assigned, dict) else "",
                "VRF": nested(address, "vrf"),
                "Status": nested(address, "status"),
                "Description": address.get("description", ""),
            }
        )
    return rows


def cable_side(cable: dict[str, Any], side: str) -> str:
    terminations = cable.get(f"{side}_terminations") or []
    return endpoint_ref(terminations)


def build_cables(cables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for cable in cables:
        rows.append(
            {
                "Cable ID": cable.get("id", ""),
                "Label": cable.get("label", ""),
                "Side A": cable_side(cable, "a"),
                "Side B": cable_side(cable, "b"),
                "Type": nested(cable, "type"),
                "Status": nested(cable, "status"),
            }
        )
    return rows


def autosize_workbook(writer: Any, helpers: dict[str, Any]) -> None:
    font = helpers["Font"](bold=True, color="FFFFFF")
    fill = helpers["PatternFill"](start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    get_column_letter = helpers["get_column_letter"]

    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = font
            cell.fill = fill

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def parse_args() -> argparse.Namespace:
    project_defaults = load_project_netbox_defaults()
    default_url = os.getenv("NETBOX_URL") or project_defaults.get("url") or "http://localhost:8000"
    default_token = os.getenv("NETBOX_TOKEN") or os.getenv("NETBOX_API_TOKEN") or project_defaults.get("token")

    parser = argparse.ArgumentParser(
        description="Export NetBox data to an Excel master workbook.",
    )
    parser.add_argument("--netbox-url", default=default_url, help=f"NetBox URL. Default: {default_url}")
    parser.add_argument("--token", default=default_token, help="NetBox API token. Can use NETBOX_TOKEN.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help=f"Excel output path. Default: {DEFAULT_OUTPUT}")
    parser.add_argument("--timeout", type=int, default=30, help="HTTP timeout in seconds.")
    parser.add_argument("--no-verify-tls", action="store_true", help="Disable TLS certificate verification.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.token:
        print(
            "NetBox API token is missing. Set NETBOX_TOKEN or pass --token.",
            file=sys.stderr,
        )
        return 2

    pd, requests, openpyxl_helpers = import_dependencies()

    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    headers = build_headers(args.token)
    verify_tls = not args.no_verify_tls

    endpoints = {
        "devices": "dcim/devices/",
        "vlans": "ipam/vlans/",
        "prefixes": "ipam/prefixes/",
        "interfaces": "dcim/interfaces/",
        "ip_addresses": "ipam/ip-addresses/",
        "cables": "dcim/cables/",
    }

    print(f"Exporting NetBox data from {args.netbox_url} ...")
    with requests.Session() as session:
        raw_data = {
            name: fetch_netbox_data(
                session=session,
                base_url=args.netbox_url,
                endpoint=endpoint,
                headers=headers,
                timeout=args.timeout,
                verify_tls=verify_tls,
            )
            for name, endpoint in endpoints.items()
        }

    sheets = {
        "Inventory": build_inventory(raw_data["devices"]),
        "IP_Seg": build_vlans(raw_data["vlans"]),
        "Prefixes": build_prefixes(raw_data["prefixes"]),
        "Maport": build_ports(raw_data["interfaces"]),
        "IP_Addresses": build_ip_addresses(raw_data["ip_addresses"]),
        "Cables": build_cables(raw_data["cables"]),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, rows in sheets.items():
            pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
        autosize_workbook(writer, openpyxl_helpers)

    print(f"Export completed: {output}")
    for sheet_name, rows in sheets.items():
        print(f"- {sheet_name}: {len(rows)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
